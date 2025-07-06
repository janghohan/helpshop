import openpyxl
import sys
import traceback


# def change_xlsx(input_file):
#     # 원본 엑셀 파일 경로
#     # 엑셀 파일 열기

#     wb = openpyxl.load_workbook(input_file)

#     ws = wb.active  # 첫 번째 시트 선택
#     ws.delete_rows(1) 
#     # 첫 번째 시트의 이름 변경
    
#     wb.active.title = "발송처리"

#     # 변경된 엑셀 파일 저장
#     wb.save(input_file)

#     # print(f"시트 이름이 '{wb.active.title}'로 변경되었습니다.")


# if __name__ == "__main__":
#         change_xlsx(sys.argv[1])

def change_xlsx(input_file):
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        ws.delete_rows(1)
        wb.active.title = "발송처리"
        wb.save(input_file)
    except Exception:
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    change_xlsx(sys.argv[1])


# import openpyxl

# # 원본 엑셀 파일 경로
# input_file = "test.xlsx"
# output_file = "renamed.xlsx"

# # 엑셀 파일 열기
# wb = openpyxl.load_workbook(input_file)

# # 첫 번째 시트의 이름 변경
# wb.active.title = "발송처리"

# # 변경된 엑셀 파일 저장
# wb.save(output_file)

# print(f"시트 이름이 '{wb.active.title}'로 변경되었습니다.")